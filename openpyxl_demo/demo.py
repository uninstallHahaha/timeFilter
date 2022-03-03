from openpyxl import load_workbook
import gc
import pandas as pd
import time
from pandas import DataFrame, Series

# 结论: 读取数据慢是慢在将数据转为dataframe, 只要数据过dataframe就逃不掉读取慢
if __name__ == '__main__':
    time_start = time.time()

    # xls -> max: 5.3%,  avg : 4.5%, about 377MB
    # xlsx -> max: 3.2%,  avg : 2.7%, about 377MB
    # pd.read_excel('test.xlsx', engine='openpyxl', storage_options={"read_only": True})
    # pd.read_excel('test.xlsx', sheet_name=None)
    # pd.DataFrame(load_workbook('test.xlsx', read_only=True))

    # mem : 1.7%, decrease 65%, about 130MB
    lw = load_workbook('test.xlsx', read_only=True)
    sns = lw.sheetnames
    # TODO: 使用此方法每次只加载一个sheet的方法使得原程序存在内存方面的优化空间
    # TODO: 1. 先对筛选条件排序
    # TODO: 2. 然后每次加载一个sheet, 获取其上下界
    # TODO: 3. 根据当前sheet的上下界完成范围内的筛选
    # TODO: 4. sheet用一个加载一个, 加载下一个之前把上一个清理掉, 筛选条件用完即删除
    # TODO: 5. 预估理论上内存占用将下降 18%~37%
    # 先用workbook把整表都出来, 然后单个65535条数据的sheet转换为DataFrame
    # max: 2%, avg 2%, decrease 37.5%, about 154MB
    s1 = DataFrame(lw[sns[0]].values)
    # 手动释放内存
    # del lw
    # del s1
    gc.collect()
    time.sleep(30)
    time_end = time.time()
    print('time cost', time_end - time_start, 's')
