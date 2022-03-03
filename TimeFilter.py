import pandas as pd
from pandas import DataFrame, Series
import numpy as np
from datetime import *
import datetime
import json
import time as tim
from tqdm import tqdm
import os
import sys
import logging
from openpyxl import load_workbook
import gc

from utils.sorts import Sorts
from utils.tools import check_mono, update_df_col, sliceList


class Datagram:
    # 1. 静止, 2.行走, 3.吃草
    def __init__(self):
        self.static_datagram = DataFrame()
        self.walk_datagram = DataFrame()
        self.eat_datagram = DataFrame()

    def __setitem__(self, key, value):
        self.key = value


# json file to obj
def json_reader(path):
    print('>>> reading configuration...')
    f = open(path, encoding='utf-8')
    obj = json.load(f)
    print('>>> read configuration over')
    return obj


# check the time whether is legal
def check_time_legal(hour: int, minite: int, second: int) -> (bool, str):
    if hour >= 24 or minite >= 60 or second >= 60:
        return False, "[error]time not legal"
    return True, ''


# str to datetime.time
def time_parser(value: str) -> (time, str):
    value = do_trim(value)
    if value == '':
        return None, ' [error] null string '
    mill_second = 0
    # shortcut-6 format
    if len(value) == 6:
        hour = int(value[:2])
        minute = int(value[2:4])
        second = int(value[4:])
    elif len(value) == 5:
        hour = int(value[:1])
        minute = int(value[1:3])
        second = int(value[3:])
    else:
        # normal-colon format
        ts = value.split(':')
        if len(ts) < 3:
            return None, f' [error] 时间格式错误 '
        if len(ts[2].split('.')) > 1:
            mill_second = int(ts[2].split('.')[1])
        hour = int(ts[0])
        minute = int(ts[1])
        second = int(ts[2].split('.')[0])
    # check legal
    c_res, reason = check_time_legal(hour, minute, second)
    if not c_res:
        return None, reason
    # res time
    refer_time = time(hour, minute, second, mill_second)
    return refer_time, ' [normal] '


# trim spaces of str
def do_trim(val: str) -> str:
    res = ''
    for i in range(len(val)):
        res += val[i] if val[i] != ' ' else ''
    return res


# check time delta
def check_time_delta(val1: datetime.time, val2: datetime.time, condition: int) -> bool:
    t1 = datetime.datetime(year=1997, month=3, day=7, hour=val1.hour, minute=val1.minute, second=val1.second)
    t2 = datetime.datetime(year=1997, month=3, day=7, hour=val2.hour, minute=val2.minute, second=val2.second)
    return abs((t2 - t1).seconds) > condition


# pre local the correct sheet
# return : shrinked dataFrame & start index of list
def shrink_dataset(origin: list, val_1: time, val_2: time, column_name='time') -> (DataFrame, int):
    # s: DataFrame
    res = None
    res_index = 0
    for index, s in enumerate(origin):
        col = s[column_name]
        first, _ = time_parser(col.iloc[0])
        last, _ = time_parser(col.iloc[-1])
        if first is None or last is None:
            continue
        min_val = first if first < last else last
        max_val = last if first < last else first
        # range beyond append
        if res is not None:
            if max_val >= val_2:
                return pd.concat([res, s], axis=0), res_index
            else:
                res = pd.concat([res, s], axis=0)
                continue
        # range content
        if min_val <= val_1 and max_val >= val_2:
            res_index = index
            return s, res_index
        # range beyond
        if min_val <= val_1 <= max_val < val_2:
            if res is None:
                res = s
                res_index = index
    return None, res_index


# 二分查找
# 返回筛选结果和上下界索引
def divide_find(fill_data: DataFrame, val1: time, val2: time, column_name='time') -> (DataFrame, int, int):
    m = map(time_parser, fill_data[column_name])
    time_column_list = [val[0] for val in list(m)]
    start = 0
    end = len(time_column_list) - 1
    loc = -1
    if time_column_list[start] > val2 or time_column_list[end] < val1:
        return DataFrame(columns=fill_data.columns), -1, -1
    while end > start:
        mid = (end - start) // 2 + start
        if start == mid and not (val1 <= time_column_list[mid] <= val2):
            break
        if val1 <= time_column_list[mid] <= val2:
            loc = mid
            break
        if time_column_list[mid] > val2:
            end = mid
            continue
        if time_column_list[mid] < val1:
            start = mid
            continue
    # not exist , return directly
    if loc == -1:
        return DataFrame(columns=fill_data.columns), -1, -1
    # step by step to found the edge of limitation
    xfound = False
    yfound = False
    x = loc
    y = loc
    while (not xfound) or (not yfound):
        if not xfound:
            if x - 1 >= 0 and val1 <= time_column_list[x - 1] <= val2:
                x -= 1
            else:
                xfound = True
        if not yfound:
            if y + 1 < len(time_column_list) and val1 <= time_column_list[y + 1] <= val2:
                y += 1
            else:
                yfound = True
    return fill_data.iloc[x:y + 1], x, y


# equal 相等
# lt 小于
# gt 大于
# lte 小于等于
# gte 大于等于
# range 范围
def mode_justifier(fill_data: DataFrame, mode: str, value1: time, value2: time, column_name='time') -> (
        DataFrame, DataFrame):
    if fill_data is None:
        return None, None, 0, 0
    filter_res = None
    remain = None
    if value1 > value2:
        return filter_res, remain, 0, 0
    if mode == 'equal':
        filter_res = fill_data[fill_data[column_name] == value1]
    if mode == 'lt':
        filter_res = fill_data[fill_data[column_name] < value1]
    if mode == 'gt':
        filter_res = fill_data[fill_data[column_name] > value1]
    if mode == 'lte':
        filter_res = fill_data[fill_data[column_name] <= value1]
    if mode == 'gte':
        max_loc = fill_data.shape[0] - 1
        if not fill_data[column_name].iloc[max_loc] < value1:
            while max_loc > 0 and fill_data[column_name].iloc[max_loc - 1] >= value1:
                max_loc -= 1
            filter_res = fill_data.iloc[max_loc:]
    if mode == 'range':
        filter_res, x, y = divide_find(fill_data=fill_data, val1=value1, val2=value2, column_name='time')
        if filter_res is not None:
            filter_res = filter_res.fillna('*')
        return filter_res, remain, x, y

        # no longer calculate remain
        # remain = fill_data.iloc[:x].append(fill_data.iloc[y + 1:])
        # remain = None

    if filter_res is not None:
        filter_res = filter_res.fillna('*')
    return filter_res, remain, 0, 0


# check historical outputs whether exists
def check_last_res(path_list):
    exi = False
    for p in path_list:
        if os.path.exists(p):
            exi = True
            break
    if exi:
        print('>>> check out there are historical outputs, new around would clean them up first, make sure you have '
              'backed them up(y/n)', end=' ')
        if input().lower() != 'y':
            print('>>> operation about')
            sys.exit(0)
        else:
            clearer(path_list)


# clean historical outputs
def clearer(path_list):
    print('>>> cleaning history output files...')
    for p in path_list:
        if os.path.exists(p):
            os.remove(p)
    print('>>> clean historical output files over')


# 根据不同的类型保存到不同的变量中
def append_datagram(res: DataFrame, time_filter: Datagram, act_type: int) -> str:
    if res is None:
        return ''
    if act_type == 1:
        time_filter.static_datagram = append_res(res, time_filter.static_datagram)
    if act_type == 2:
        time_filter.walk_datagram = append_res(res, time_filter.walk_datagram)
    if act_type == 3:
        time_filter.eat_datagram = append_res(res, time_filter.eat_datagram)
    # res.to_excel(path, index=False, header=False)
    return ''


# append filter result
def append_res(res, datagram):
    df = datagram
    if df is None or df.shape[0] == 0:
        return res
    res.index = map(lambda x: x, range(df.shape[0], df.shape[0] + res.shape[0]))
    res.columns = df.columns
    df.reset_index(drop=True)
    res.reset_index(drop=True)
    res = df.append(res, ignore_index=True)
    return res


# save to files from Datagram
def saver_all_files(dg: Datagram, paths: list):
    for attr, d in dg.__dict__.items():
        if attr == 'static_datagram':
            d.to_excel(paths[0], index=False, header=False)
        elif attr == 'walk_datagram':
            d.to_excel(paths[1], index=False, header=False)
        elif attr == 'eat_datagram':
            d.to_excel(paths[2], index=False, header=False)
        else:
            pass


# saver for other modes
def saver(ret, file_name='result'):
    path = './' + file_name + '-' + str(round(tim.time()))[-1] + '.xlsx'
    ret.to_excel(path)
    print('保存成功: ' + path)


# rebuild outputs(replace *)
def re_format_table(datagram: Datagram, old_val='*', new_val=''):
    print('>>> rebuilding output...')
    with tqdm(total=3, leave=True) as bar:
        bar.set_description(" rebuilding ")
        for attr in list(datagram.__dict__.keys()):
            datagram.__dict__[attr] = datagram.__dict__[attr].replace(old_val, new_val)
            bar.update(1)

    return


# set logging
def set_logging(action='', path='log.txt', log_name='log',
                formatter='%(asctime)s - %(name)s - %(levelname)s - %(message)s'):
    logger = logging.getLogger(log_name)
    logger.setLevel(level=logging.INFO)
    # create log dir
    dir_path = 'logs/' + log_name
    os.makedirs(dir_path, exist_ok=True)
    # get full file name
    now = datetime.datetime.now().strftime('%Y-%m-%d-')
    path = 'logs/' + log_name + '/' + now + action + path
    handler = logging.FileHandler(path)
    handler.setLevel(logging.INFO)
    formatter = logging.Formatter(formatter)
    handler.setFormatter(formatter)
    logger.addHandler(handler)
    logger.info(f'')
    return logger


def finish_filter(pbar: tqdm):
    pass


# check whether could get limitation of df
def can_limitation(df: DataFrame, col: int) -> bool:
    if len(df.columns) <= col:
        return False
    return True


# extract the min time and max time of df
def extract_limitation(df: DataFrame, col: int) -> (time, time):
    # if len(df.columns) <= col:
    #     return None, None
    min_time, _ = time_parser(df.iloc[0].iloc[col])
    max_time, _ = time_parser(df.iloc[-1].iloc[col])
    return min_time, max_time


if __name__ == '__main__':
    # log logger
    logger = set_logging()
    logger_static = set_logging(action='static', log_name='static')
    logger_walk = set_logging(action='walk', log_name='walk')
    logger_eat = set_logging(action='eat', log_name='eat')
    loggers = [logger_static, logger_walk, logger_eat]
    # err logger
    logger_err = set_logging(path='log_error.txt', log_name='log_error')
    # config
    settings = json_reader('./settings.json')
    # file name
    base_name = settings["file_name"].split('.')[0]
    # file names of output files
    paths = [base_name + '-静止.xlsx', base_name + '-行走.xlsx', base_name + '-吃草.xlsx']
    # check last output files
    check_last_res(paths)

    print('>>> reading data...')
    time_start = tim.time()

    # amount of all data
    total = 0
    # all the data of sheets'
    all_sheet_data = None
    # sheet data list
    # speed mode : table data of sheets
    # mem mode : name of sheets
    sheet_list = []
    # amount of sheet
    num_sheet = 0
    # table for next filter
    tmpTable = None
    # the sheet which is pointed now
    sheet_loc = 0

    if settings["mem_save"] == 1 and str(settings["file_name"]).endswith('xls'):
        print(">>> mem first mode only support 'xlsx' files, convert origin file to 'xlsx' before run this program")
        sys.exit(0)

    if settings["mem_save"] == 1:
        all_sheet_data = load_workbook(settings['file_name'], read_only=True)
        sheet_list = all_sheet_data.sheetnames
        for s in sheet_list:
            total += all_sheet_data[s].max_row
        # get the first sheet which is not empty
        while (not update_df_col(tmpTable, loc=settings['time_column_num'])) and sheet_loc != -1:
            tmpTable = DataFrame(all_sheet_data[sheet_list[sheet_loc]].values)
            sheet_loc = sheet_loc + 1 if sheet_loc + 1 < len(sheet_list) else -1
        num_sheet = len(sheet_list)
    else:
        all_sheet_data = pd.read_excel(settings['file_name'], header=None, sheet_name=None)
        num_sheet = len(all_sheet_data)

    # fillTable = None

    # prepare data for speed mode
    if settings['mem_save'] != 1:
        for sheet_name, data in all_sheet_data.items():
            cols = ['-'] * data.shape[1]
            # continue if the sheet is not content time column
            if len(cols) < int(settings['time_column_num']):
                continue
            # set the filter column's name to 'time'
            cols[int(settings['time_column_num']) - 1] = 'time'
            data.columns = cols
            data.index = map(lambda x: x, range(total, total + data.shape[0]))

            # append to fillTable
            # fillTable = pd.concat([fillTable, data], axis=0)
            # append to list
            sheet_list.append(data)
            total += len(data)

        # slice sheet_list tobe more delicate
        sheet_list = sliceList(sheet_list, settings['slice_bucket'])
        del all_sheet_data
        gc.collect()

    time_end = tim.time()
    print(
        f'>>> reading data over, consist of {total} records, {num_sheet} sheets, spend {round(time_end - time_start, 2)} s')

    # multi-range-filter
    if settings["mode"] == 'range':
        # conditions
        range_data = pd.read_excel(settings['range_file_name'], header=None)
        range_data_len = len(range_data)
        # where to save filter results
        datagram = Datagram()
        print(f'>>> Start Range mode({"mem first" if settings["mem_save"] == True else "speed first"})...')

        # check ascent and sort it
        if settings['mem_save'] == 1:
            check_res: bool = check_mono(range_data)
            if not check_res:
                print(f">>> start sort conditions as checked out start conditions are not ascent...")
                print(f'>>> note: if start conditions are ascent, it would skip this step for more efficiency')
                Sorts.fast_sort(range_data, 0, len(range_data) - 1)

        # iter all of conditions
        with tqdm(total=range_data_len, leave=True) as pbar:
            pbar.set_description(" Filtering ")
            filter_done: bool = False
            for index, row in range_data.iterrows():
                if filter_done:
                    pbar.update(pbar.total)
                    break
                pbar.update(1)
                # less than 3 cell's row would be ignored
                if len(row) < 3:
                    continue
                # the begin time of iter
                t_start = tim.time()
                # conditions
                try:
                    if type(row[0]) != time:
                        val1, reason1 = time_parser(str(int(row[0])))
                    else:
                        val1 = row[0]
                        reason1 = ''
                    if type(row[1]) != time:
                        val2, reason2 = time_parser(str(int(row[1])))
                    else:
                        val2 = row[1]
                        reason2 = ''
                    filter_type = int(row[2])
                    # range condition is None, continue directly
                    if val1 is None or val2 is None:
                        if val1 is None:
                            logger_err.info(f'第 {index + 1} 条筛选条件解析失败, 自动跳过, 因为起始时间 {reason1}')
                            continue
                        if val2 is None:
                            logger_err.info(f'第 {index + 1} 条筛选条件解析失败, 自动跳过, 因为截止时间 {reason2}')
                            continue
                    # type out of condition
                    if filter_type < 1 or filter_type > 3:
                        logger_err.info(f'第 {index + 1} 条筛选条件解析失败, 自动跳过, 是因为第三列的类型没有按照标准来')
                        continue
                    if filter_type == 2 or filter_type == 3:
                        if check_time_delta(val1, val2, settings['error_range']):
                            logger_err.info(f'第 {index + 1} 条筛选条件时间范围异常，超过了{settings["error_range"]}s, 自动跳过')
                            continue
                except:
                    logger_err.info(f'第 {index + 1} 条筛选条件解析失败, 自动跳过')
                    continue

                # prepare table to be filtered
                start_index = 0
                if settings['mem_save'] != 1:
                    # shrink the range of sheet from sheet's list
                    tmpTable, start_index = shrink_dataset(sheet_list, val1, val2)
                else:
                    # cur sheet's limitation
                    min_limit, max_limit = extract_limitation(df=tmpTable, col=settings['time_column_num'] - 1)
                    # target current sheet
                    if min_limit <= val1 <= val2 <= max_limit or min_limit <= val1 <= max_limit <= val2 or val1 <= min_limit <= val2 <= max_limit or val1 <= min_limit <= max_limit <= val2:
                        pass
                    # beyond current sheet, find next correct shit
                    elif max_limit < val1 < val2:
                        if sheet_loc == -1:
                            filter_done = True
                            continue
                        tmpTable = DataFrame(all_sheet_data[sheet_list[sheet_loc]].values)
                        sheet_loc = sheet_loc + 1 if sheet_loc + 1 < len(sheet_list) else -1
                        gc.collect()

                        while not update_df_col(tmpTable, loc=settings['time_column_num']) or not (
                                can_limitation(tmpTable, col=settings['time_column_num']) or not (
                                extract_limitation(df=tmpTable, col=settings['time_column_num'] - 1)[
                                    0] <= val1 <= val2 <=
                                extract_limitation(df=tmpTable, col=settings['time_column_num'] - 1)[1] or
                                extract_limitation(df=tmpTable, col=settings['time_column_num'] - 1)[0] <= val1 <=
                                extract_limitation(df=tmpTable, col=settings['time_column_num'] - 1)[1] <= val2 or
                                val1 <= extract_limitation(df=tmpTable, col=settings['time_column_num'] - 1)[
                                    0] <= val2 <= extract_limitation(df=tmpTable, col=settings['time_column_num'] - 1)[
                                    1] or
                                val1 <= extract_limitation(df=tmpTable, col=settings['time_column_num'] - 1)[0] <=
                                extract_limitation(df=tmpTable, col=settings['time_column_num'] - 1)[1] <= val2
                        )):
                            if sheet_loc == -1:
                                tmpTable = None
                                break
                            tmpTable = DataFrame(all_sheet_data[sheet_list[sheet_loc]].values)
                            sheet_loc = sheet_loc + 1 if sheet_loc + 1 < len(sheet_list) else -1
                            # if can_limitation(tmpTable, col=settings['time_column_num']):
                            #     min_limit, max_limit =
                            #     if :
                            #         break
                            gc.collect()
                        if tmpTable is None:
                            filter_done = True
                            continue
                        # gc.collect()
                    # ride current shit and next shit
                    # elif min_limit <= val1 <= max_limit < val2:
                    #     # filter from last sheet
                    #     res, _ = mode_justifier(tmpTable, mode='gte', value1=val1)
                    #     # update fill_data of classes
                    #     append_datagram(res, datagram, filter_type)
                    #
                    #     # load next sheet
                    #     if sheet_loc == -1:
                    #         filter_done = True
                    #     tmpTable = DataFrame(all_sheet_data[sheet_list[sheet_loc]].values)
                    #     while (not update_df_col(tmpTable, loc=settings['time_column_num'])) and sheet_loc != -1:
                    #         tmpTable = DataFrame(all_sheet_data[sheet_list[sheet_loc]].values)
                    #         sheet_loc = sheet_loc + 1 if sheet_loc + 1 < len(sheet_list) else -1
                    #     if sheet_loc == -1:
                    #         filter_done = True
                    #     # update range
                    #     val1 = tmpTable.iloc[0].iloc[settings['time_column_num'] - 1]

                # filter from tmpTable
                if tmpTable is None and not filter_done:
                    t_end = tim.time()
                    # write total log
                    logger.info(
                        f"{'(mem mode)' if settings['mem_save'] == 1 else '(sp mode)'} 第 {index + 1} 号筛选条件(类别{int(row[2])}, 条件范围({val1.strftime('%H:%M:%S')}~{val2.strftime('%H:%M:%S')})), 耗时 {round(t_end - t_start, 2)} 秒, 筛选结果为 0 条数据")
                    # write action log
                    loggers[int(row[2]) - 1].info(
                        f"{'(mem mode)' if settings['mem_save'] == 1 else '(sp mode)'} 第 {index + 1} 号筛选条件(类别{int(row[2])}), 条件范围({val1.strftime('%H:%M:%S')}~{val2.strftime('%H:%M:%S')}), 耗时 {round(t_end - t_start, 2)} 秒, 筛选结果为 0 条数据")
                    continue
                # get the record which is in the range
                ret, _, x, y = mode_justifier(tmpTable, mode='range', value1=val1, value2=val2)
                t_end = tim.time()
                # write total log
                logger.info(
                    f"第 {index + 1} 号筛选条件(类别{int(row[2])})耗时 {round(t_end - t_start, 2)} 秒, "
                    f"条件范围({val1.strftime('%H:%M:%S')}~{val2.strftime('%H:%M:%S')}), "
                    f"在原表中命中的范围为({start_index * settings['slice_bucket'] + x + 1}~{start_index * settings['slice_bucket'] + y + 1}), "
                    f"筛选结果为 {ret.shape[0] if type(ret) == DataFrame else 0} 条数据")
                # write action log
                loggers[int(row[2]) - 1].info(
                    f"第 {index + 1} 号筛选条件(类别{int(row[2])})耗时 {round(t_end - t_start, 2)} 秒, "
                    f"条件范围({val1.strftime('%H:%M:%S')}~{val2.strftime('%H:%M:%S')}), "
                    f"在原表中命中的范围为({start_index * settings['slice_bucket'] + x + 1}~{start_index * settings['slice_bucket'] + y + 1}), "
                    f"筛选结果为 {ret.shape[0] if type(ret) == DataFrame else 0} 条数据")

                # update fillTable to remain_data
                # fillTable = remain_data

                # update fill_data of classes
                append_datagram(ret, datagram, filter_type)

            print('>>> filter working is over')

        # rebuild datagram(replace * to '')
        re_format_table(datagram)
        # output all result data to files
        saver_all_files(datagram, paths)
        print('>>> rebuilding over, all is done')

    # other mode
    # else:
    # other-mode
    # ret = mode_justifier(fillTable, mode=settings['mode'], value1=time_parser(settings['value1']),
    #                      value2=time_parser(settings['value2']))
    # saver(ret)
